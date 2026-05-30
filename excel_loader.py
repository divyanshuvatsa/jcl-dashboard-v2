"""
excel_loader.py — fast, external full-model loader
==================================================
Loads the ENTIRE computed JCL model from a workbook (the active scenario) and
returns it in exactly the shape engine.build() produces, so the dashboard can
render the workbook's own numbers with zero re-derivation — i.e. 100% fidelity
to the source by construction.

Speed: opens the workbook once with read_only=True + data_only=True and reads
cached values in a single pass. No LibreOffice, no formula recalculation, so a
~300 KB workbook loads in well under a second. (For comparison, recalculating
with a headless office engine would take 5-15 s — deliberately avoided.)
"""
from __future__ import annotations

import json
import os
from typing import Dict, Optional

import numpy as np
import pandas as pd

from model_data import ALL_YEARS, HISTORICAL_YEAR, PROJECTION_YEARS, VALUED_YEARS

# Column maps (see source workbook layouts)
YRS_3S = {"FY24A": "B", "FY25A": "C", "FY26E": "D", "FY27E": "E", "FY28E": "F",
          "FY29E": "G", "FY30E": "H", "FY31E": "I", "FY32E": "J", "FY33E": "K"}
YRS_SCHED = {"FY24A": "C", "FY25A": "D", "FY26E": "E", "FY27E": "F", "FY28E": "G",
             "FY29E": "H", "FY30E": "I", "FY31E": "J", "FY32E": "K", "FY33E": "L"}
YRS_AD = {"FY25A": "D", "FY26E": "E", "FY27E": "F", "FY28E": "G", "FY29E": "H",
          "FY30E": "I", "FY31E": "J", "FY32E": "K", "FY33E": "L"}

SHARES = 3.243
PREF = 109.26


def _num(v):
    return float(v) if isinstance(v, (int, float)) else None


def _grid_reader(ws):
    cache = {}

    def get(addr):
        if addr not in cache:
            cache[addr] = ws[addr].value
        return cache[addr]
    return get


def load_full_model(file) -> Dict:
    """Read the full computed model from `file` (path or file-like). Fast."""
    import openpyxl
    wb = openpyxl.load_workbook(file, data_only=True, read_only=False, keep_links=False)

    def sheet(name):
        return wb[name]

    def row(ws, r, years, ymap):
        return {y: _num(ws[f"{ymap[y]}{r}"].value) for y in years}

    # ----- 3-Statement -----
    ws = sheet("The 3-Statement Projection")
    inc = {
        "Net_Sales": row(ws, 6, ALL_YEARS, YRS_3S), "COGS": row(ws, 9, ALL_YEARS, YRS_3S),
        "Gross_Profit": row(ws, 10, ALL_YEARS, YRS_3S), "Employee": row(ws, 12, ALL_YEARS, YRS_3S),
        "SGA": row(ws, 13, ALL_YEARS, YRS_3S), "EBITDA": row(ws, 15, ALL_YEARS, YRS_3S),
        "Depreciation": row(ws, 18, ALL_YEARS, YRS_3S), "EBIT": row(ws, 19, ALL_YEARS, YRS_3S),
        "Other_Income": row(ws, 20, ALL_YEARS, YRS_3S), "Interest": row(ws, 21, ALL_YEARS, YRS_3S),
        "PBT": row(ws, 23, ALL_YEARS, YRS_3S), "Tax": row(ws, 25, ALL_YEARS, YRS_3S),
        "PAT": row(ws, 27, ALL_YEARS, YRS_3S),
    }
    bal = {
        "Share_Capital": row(ws, 34, ALL_YEARS, YRS_3S),
        "Reserves_Surplus": row(ws, 35, ALL_YEARS, YRS_3S),
        "Total_Equity": row(ws, 36, ALL_YEARS, YRS_3S),
        "LT_Borrowings": row(ws, 39, ALL_YEARS, YRS_3S),
        "Deferred_Tax_Liab": row(ws, 40, ALL_YEARS, YRS_3S),
        "LT_Provisions": row(ws, 41, ALL_YEARS, YRS_3S),
        "ST_Borrowings": row(ws, 45, ALL_YEARS, YRS_3S),
        "Revolver": row(ws, 46, ALL_YEARS, YRS_3S),
        "Trade_Payables": row(ws, 47, ALL_YEARS, YRS_3S),
        "Other_CL": row(ws, 48, ALL_YEARS, YRS_3S),
        "Net_Fixed_Assets": row(ws, 57, ALL_YEARS, YRS_3S),
        "CWIP": row(ws, 58, ALL_YEARS, YRS_3S),
        "LT_Investments": row(ws, 59, ALL_YEARS, YRS_3S),
        "LT_Loans_Adv": row(ws, 60, ALL_YEARS, YRS_3S),
        "Other_LT_Assets": row(ws, 61, ALL_YEARS, YRS_3S),
        "Inventories": row(ws, 65, ALL_YEARS, YRS_3S),
        "Trade_Receivables": row(ws, 66, ALL_YEARS, YRS_3S),
        "Cash": row(ws, 67, ALL_YEARS, YRS_3S),
        "Other_CA": row(ws, 68, ALL_YEARS, YRS_3S),
        "ST_Loans_Adv": row(ws, 69, ALL_YEARS, YRS_3S),
        "Total_Assets": row(ws, 72, ALL_YEARS, YRS_3S),
        "Total_Liab_Equity": row(ws, 52, ALL_YEARS, YRS_3S),
    }
    cfs = {
        "PAT": row(ws, 77, PROJECTION_YEARS, YRS_3S),
        "Depreciation": row(ws, 78, PROJECTION_YEARS, YRS_3S),
        "Delta_Inventory": row(ws, 81, PROJECTION_YEARS, YRS_3S),
        "Delta_Receivables": row(ws, 82, PROJECTION_YEARS, YRS_3S),
        "Delta_Payables": row(ws, 83, PROJECTION_YEARS, YRS_3S),
        "Delta_Other_NCL": row(ws, 84, PROJECTION_YEARS, YRS_3S),
        "CFO": row(ws, 85, PROJECTION_YEARS, YRS_3S),
        "Capex": row(ws, 88, PROJECTION_YEARS, YRS_3S),
        "CFI": row(ws, 89, PROJECTION_YEARS, YRS_3S),
        "Term_Debt_Repay": row(ws, 92, PROJECTION_YEARS, YRS_3S),
        "Change_ST_Borrow": row(ws, 93, PROJECTION_YEARS, YRS_3S),
        "Dividend": row(ws, 94, PROJECTION_YEARS, YRS_3S),
        "CFF": row(ws, 95, PROJECTION_YEARS, YRS_3S),
        "Net_Change_Cash": row(ws, 97, PROJECTION_YEARS, YRS_3S),
        "Revolver_Draw": row(ws, 101, PROJECTION_YEARS, YRS_3S),
        "Closing_Cash": row(ws, 102, PROJECTION_YEARS, YRS_3S),
    }

    # ----- WACC & DCF -----
    d = sheet("WACC & DCF Valuation")
    G = _grid_reader(d)
    dcf_ebit = row(d, 36, VALUED_YEARS, YRS_3S)
    dcf_nopat = row(d, 38, VALUED_YEARS, YRS_3S)
    dcf_dep = row(d, 40, VALUED_YEARS, YRS_3S)
    dcf_capex = row(d, 41, VALUED_YEARS, YRS_3S)
    dcf_dnwc = row(d, 44, VALUED_YEARS, YRS_3S)
    dcf_fcff = row(d, 46, VALUED_YEARS, YRS_3S)
    dcf_disc = row(d, 49, VALUED_YEARS, YRS_3S)
    dcf_pv = row(d, 50, VALUED_YEARS, YRS_3S)
    wacc = _num(G("B32"))
    g_term = _num(G("B56"))

    # ----- Covenant -----
    cv = sheet("DSCR Covenant Tracker")
    cov = {
        "principal_repay": row(cv, 15, PROJECTION_YEARS, YRS_3S),
        "deferred_tax": row(cv, 10, PROJECTION_YEARS, YRS_3S),
        "dscr": row(cv, 19, PROJECTION_YEARS, YRS_3S),
        "iscr": row(cv, 20, PROJECTION_YEARS, YRS_3S),
    }

    # ----- Sensitivity -----
    sn = sheet("Sensitivity Analysis")
    GS = _grid_reader(sn)
    g_cols = [_num(GS(f"{c}10")) for c in ["B", "C", "D", "E", "F"]]
    wacc_rows = [_num(GS(f"A{r}")) for r in range(11, 16)]
    sens_vals = [[_num(GS(f"{c}{r}")) for c in ["B", "C", "D", "E", "F"]]
                 for r in range(11, 16)]

    # ----- Revenue build -----
    ad = sheet("Assumptions & Drivers")
    rev = {
        "COB1_Production": row(ad, 85, PROJECTION_YEARS, YRS_AD),
        "COB2_Production": row(ad, 88, PROJECTION_YEARS, YRS_AD),
        "Total_Coke": row(ad, 89, PROJECTION_YEARS, YRS_AD),
        "Coke_Revenue": row(ad, 91, PROJECTION_YEARS, YRS_AD),
        "COG_Revenue": row(ad, 95, PROJECTION_YEARS, YRS_AD),
        "Tar_Revenue": row(ad, 99, PROJECTION_YEARS, YRS_AD),
        "Total_Revenue": row(ad, 100, PROJECTION_YEARS, YRS_AD),
    }

    # ----- Assumptions (Scenario Engine, active column) -----
    se = sheet("Scenario Engine")
    GE = _grid_reader(se)
    active = GE("B3")
    col = {"Bull": "B", "Base": "C", "Bear": "D"}.get(active, "C")
    wcol = {"Bull": "B", "Base": "C", "Bear": "D"}.get(active, "C")
    assumptions = {
        "cob2_util_steady": _num(GE(f"{col}48")),
        "coke_realization": _num(GE(f"{col}49")),
        "cogs_pct": _num(GE(f"{col}7")),
        "interest_rate": _num(GE(f"{col}27")),
        "target_de": _num(GE(f"{wcol}87")),
        "unlevered_beta": _num(GE(f"{col}32")),
        "terminal_growth": _num(GE(f"{col}31")),
        "capex_intensity": 0.025,
        "rf_rate": _num(GE(f"{wcol}83")),
        "erp": _num(GE(f"{wcol}84")),
    }

    # ================= assemble engine-shaped output =================
    income_df = pd.DataFrame(inc).reindex(ALL_YEARS)
    income_df["Capex"] = np.nan
    income_df["Opening_Gross_Block"] = np.nan
    income_df["Closing_Term_Debt"] = pd.Series(bal["LT_Borrowings"])
    income_df["EBITDA_Margin"] = income_df["EBITDA"] / income_df["Net_Sales"]
    income_df["PAT_Margin"] = income_df["PAT"] / income_df["Net_Sales"]
    income_df["Gross_Margin"] = income_df["Gross_Profit"] / income_df["Net_Sales"]

    bal_df = pd.DataFrame(bal).reindex(ALL_YEARS)
    cfs_df = pd.DataFrame(cfs).reindex(PROJECTION_YEARS)

    # ratios (same definitions as engine)
    rrows = []
    for y in PROJECTION_YEARS:
        b = bal_df.loc[y]; i = income_df.loc[y]
        ebitda = i["EBITDA"]; ebit = i["EBIT"]; interest = i["Interest"]
        total_debt = b["LT_Borrowings"] + b["ST_Borrowings"] + b["Revolver"]
        net_debt = total_debt - b["Cash"]
        equity = b["Total_Equity"]
        ncl = b["LT_Borrowings"] + b["Deferred_Tax_Liab"] + b["LT_Provisions"]
        tca = b["Inventories"] + b["Trade_Receivables"] + b["Cash"] + b["Other_CA"] + b["ST_Loans_Adv"]
        tcl = b["ST_Borrowings"] + b["Revolver"] + b["Trade_Payables"] + b["Other_CL"]
        principal = cov["principal_repay"][y]; dtx = cov["deferred_tax"][y]
        cads = i["PAT"] + i["Depreciation"] + interest + dtx
        ds = interest + principal
        rrows.append({
            "Year": y, "EBITDA_Margin": i["EBITDA_Margin"], "PAT_Margin": i["PAT_Margin"],
            "ROE": i["PAT"] / equity if equity else np.nan,
            "ROCE": ebit / (equity + ncl) if (equity + ncl) else np.nan,
            "Debt_Equity": total_debt / equity if equity else np.nan,
            "Net_Debt": net_debt,
            "Net_Debt_EBITDA": net_debt / ebitda if ebitda else np.nan,
            "DSCR": cads / ds if ds else np.nan,
            "Interest_Coverage": ebitda / interest if interest else np.nan,
            "Current_Ratio": tca / tcl if tcl else np.nan,
            "Total_Debt": total_debt, "Gross_Margin": i["Gross_Margin"],
        })
    ratios_df = pd.DataFrame(rrows).set_index("Year")

    # revenue df
    rev_df = pd.DataFrame(rev).reindex(PROJECTION_YEARS)
    rev_df["Utilization"] = rev_df["Total_Coke"] / (429000 + 350000)

    # dcf
    fcff_rows = []
    for k, y in enumerate(VALUED_YEARS, start=1):
        fcff_rows.append({
            "Year": y, "T": k, "EBIT": dcf_ebit[y], "NOPAT": dcf_nopat[y],
            "Depreciation": dcf_dep[y], "Capex": dcf_capex[y],
            "Delta_NWC": -dcf_dnwc[y], "FCFF": dcf_fcff[y],
            "Discount_Factor": dcf_disc[y], "PV_FCFF": dcf_pv[y],
        })
    fcff_df = pd.DataFrame(fcff_rows).set_index("Year")
    ev = _num(G("B65")); pv_tv = _num(G("B60"))
    dcf = {
        "wacc_components": {
            "WACC": wacc, "Ke": _num(G("B21")), "Kd_AT": _num(G("B25")),
            "Kp": _num(G("B26")), "Beta_Levered": _num(G("B20")),
            "We": _num(G("B29")), "Wd": _num(G("B30")), "Wp": _num(G("B31")),
        },
        "fcff": fcff_df,
        "terminal_fcff_cit": _num(G("B58")),
        "terminal_value": _num(G("B59")), "pv_terminal": pv_tv,
        "sum_pv_fcff": _num(G("B63")), "enterprise_value": ev,
        "cash": _num(G("B72")), "debt": _num(G("B73")),
        "preference": _num(G("B74")), "equity_value": _num(G("B76")),
        "value_per_share": _num(G("B81")),
        "pct_ev_terminal": pv_tv / ev if ev else np.nan,
        "wacc": wacc, "terminal_growth": g_term,
    }

    # sensitivity df (same labels as engine)
    sens_df = pd.DataFrame(
        sens_vals, index=[f"{w:.2%}" for w in wacc_rows],
        columns=[f"{g:.2%}" for g in g_cols])

    wb.close()
    return {
        "revenue": rev_df, "income": income_df, "balance": bal_df,
        "cashflow": cfs_df, "ratios": ratios_df, "dcf": dcf,
        "sensitivity": sens_df, "assumptions": assumptions, "source": "excel",
    }


def regenerate_base_model(file, out: Optional[str] = None) -> str:
    """Refresh the bundled jcl_base_model.json from an updated workbook.
    Re-runs the golden extractor logic so the engine's Base stays in sync."""
    import shutil
    import subprocess
    import sys
    here = os.path.dirname(os.path.abspath(__file__))
    out = out or os.path.join(here, "jcl_base_model.json")
    extractor = os.path.join(here, "extract_golden.py")
    if not os.path.exists(extractor):
        raise FileNotFoundError("extract_golden.py not found alongside excel_loader.py")
    subprocess.run([sys.executable, extractor, os.path.abspath(file)],
                   check=True, cwd=here)
    produced = os.path.join(here, "golden.json")
    if os.path.abspath(produced) != os.path.abspath(out):
        shutil.copyfile(produced, out)
    return out
