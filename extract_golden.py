"""
extract_golden.py — pull the FULL computed JCL model from the Excel workbook
at full float precision, into golden.json. This is the single ground-truth
artifact used both for (a) verification and (b) engine/loader calibration.

Reads cached (data_only) values — no recalculation needed.
"""
import json
import sys
import warnings
import openpyxl

warnings.filterwarnings("ignore")

SRC = sys.argv[1] if len(sys.argv) > 1 else "JCL_Financial_Model_EXP.xlsx"

wb = openpyxl.load_workbook(SRC, data_only=True)

# Column letters for years on the 3-Statement / DCF / schedules
# 3-Statement: B=FY24A C=FY25A D=FY26E ... K=FY33E
YRS_3S = {"FY24A": "B", "FY25A": "C", "FY26E": "D", "FY27E": "E", "FY28E": "F",
          "FY29E": "G", "FY30E": "H", "FY31E": "I", "FY32E": "J", "FY33E": "K"}
# Debt / Depreciation schedules: C=FY24A D=FY25A E=FY26E ... L=FY33E
YRS_SCHED = {"FY24A": "C", "FY25A": "D", "FY26E": "E", "FY27E": "F", "FY28E": "G",
             "FY29E": "H", "FY30E": "I", "FY31E": "J", "FY32E": "K", "FY33E": "L"}
PROJ = ["FY25A", "FY26E", "FY27E", "FY28E", "FY29E", "FY30E", "FY31E", "FY32E", "FY33E"]
ALL = ["FY24A"] + PROJ


def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    return None


def row(ws, r, years, ymap):
    return {y: num(ws[f"{ymap[y]}{r}"].value) for y in years}


g = {}

# ---------------------------------------------------------------- 3-Statement
ws = wb["The 3-Statement Projection"]
g["income"] = {
    "Net_Sales":     row(ws, 6, ALL, YRS_3S),
    "COGS":          row(ws, 9, ALL, YRS_3S),
    "Gross_Profit":  row(ws, 10, ALL, YRS_3S),
    "Employee":      row(ws, 12, ALL, YRS_3S),
    "SGA":           row(ws, 13, ALL, YRS_3S),
    "EBITDA":        row(ws, 15, ALL, YRS_3S),
    "Depreciation":  row(ws, 18, ALL, YRS_3S),
    "EBIT":          row(ws, 19, ALL, YRS_3S),
    "Other_Income":  row(ws, 20, ALL, YRS_3S),
    "Interest":      row(ws, 21, ALL, YRS_3S),
    "PBT":           row(ws, 23, ALL, YRS_3S),
    "Tax":           row(ws, 25, ALL, YRS_3S),
    "PAT":           row(ws, 27, ALL, YRS_3S),
}
g["balance"] = {
    "Share_Capital":     row(ws, 34, ALL, YRS_3S),
    "Reserves_Surplus":  row(ws, 35, ALL, YRS_3S),
    "Total_Equity":      row(ws, 36, ALL, YRS_3S),
    "LT_Borrowings":     row(ws, 39, ALL, YRS_3S),
    "Deferred_Tax_Liab": row(ws, 40, ALL, YRS_3S),
    "LT_Provisions":     row(ws, 41, ALL, YRS_3S),
    "ST_Borrowings":     row(ws, 45, ALL, YRS_3S),
    "Revolver":          row(ws, 46, ALL, YRS_3S),
    "Trade_Payables":    row(ws, 47, ALL, YRS_3S),
    "Other_CL":          row(ws, 48, ALL, YRS_3S),
    "Net_Fixed_Assets":  row(ws, 57, ALL, YRS_3S),
    "CWIP":              row(ws, 58, ALL, YRS_3S),
    "LT_Investments":    row(ws, 59, ALL, YRS_3S),
    "LT_Loans_Adv":      row(ws, 60, ALL, YRS_3S),
    "Other_LT_Assets":   row(ws, 61, ALL, YRS_3S),
    "Inventories":       row(ws, 65, ALL, YRS_3S),
    "Trade_Receivables": row(ws, 66, ALL, YRS_3S),
    "Cash":              row(ws, 67, ALL, YRS_3S),
    "Other_CA":          row(ws, 68, ALL, YRS_3S),
    "ST_Loans_Adv":      row(ws, 69, ALL, YRS_3S),
    "Total_Assets":      row(ws, 72, ALL, YRS_3S),
    "Total_Liab_Equity": row(ws, 52, ALL, YRS_3S),
}
g["cashflow"] = {
    "PAT":               row(ws, 77, PROJ, YRS_3S),
    "Depreciation":      row(ws, 78, PROJ, YRS_3S),
    "Delta_Inventory":   row(ws, 81, PROJ, YRS_3S),
    "Delta_Receivables": row(ws, 82, PROJ, YRS_3S),
    "Delta_Payables":    row(ws, 83, PROJ, YRS_3S),
    "Delta_Other_NCL":   row(ws, 84, PROJ, YRS_3S),
    "CFO":               row(ws, 85, PROJ, YRS_3S),
    "Capex":             row(ws, 88, PROJ, YRS_3S),
    "CFI":               row(ws, 89, PROJ, YRS_3S),
    "Term_Debt_Repay":   row(ws, 92, PROJ, YRS_3S),
    "Change_ST_Borrow":  row(ws, 93, PROJ, YRS_3S),
    "Dividend":          row(ws, 94, PROJ, YRS_3S),
    "CFF":               row(ws, 95, PROJ, YRS_3S),
    "Net_Change_Cash":   row(ws, 97, PROJ, YRS_3S),
    "Opening_Cash":      row(ws, 99, PROJ, YRS_3S),
    "Revolver_Draw":     row(ws, 101, PROJ, YRS_3S),
    "Closing_Cash":      row(ws, 102, PROJ, YRS_3S),
}

# ---------------------------------------------------------------- WACC & DCF
ws = wb["WACC & DCF Valuation"]
YRS_DCF = YRS_3S  # same layout B..K
g["dcf"] = {
    "rf_rate":        num(ws["B7"].value),
    "erp":            num(ws["B8"].value),
    "unlevered_beta": num(ws["B9"].value),
    "marginal_tax":   num(ws["B10"].value),
    "mat_rate":       num(ws["B11"].value),
    "equity":         num(ws["B13"].value),
    "debt":           num(ws["B14"].value),
    "preference":     num(ws["B15"].value),
    "total_capital":  num(ws["B16"].value),
    "target_de":      num(ws["B17"].value),
    "beta_levered":   num(ws["B20"].value),
    "ke":             num(ws["B21"].value),
    "pretax_kd":      num(ws["B24"].value),
    "kd_at":          num(ws["B25"].value),
    "kp":             num(ws["B26"].value),
    "we":             num(ws["B29"].value),
    "wd":             num(ws["B30"].value),
    "wp":             num(ws["B31"].value),
    "wacc":           num(ws["B32"].value),
    "ebit":            row(ws, 36, PROJ, YRS_DCF),
    "tax_on_ebit":     row(ws, 37, PROJ, YRS_DCF),
    "nopat":           row(ws, 38, PROJ, YRS_DCF),
    "dep":             row(ws, 40, PROJ, YRS_DCF),
    "capex":           row(ws, 41, PROJ, YRS_DCF),
    "nwc":             row(ws, 43, ALL, YRS_DCF),
    "delta_nwc":       row(ws, 44, PROJ, YRS_DCF),
    "fcff":            row(ws, 46, PROJ, YRS_DCF),
    "discount_factor": row(ws, 49, PROJ, YRS_DCF),
    "pv_fcff":         row(ws, 50, PROJ, YRS_DCF),
    "sum_pv_fcff":     num(ws["C51"].value),
    "terminal_growth": num(ws["B56"].value),
    "final_fcff":      num(ws["B57"].value),
    "terminal_fcff_cit": num(ws["B58"].value),
    "terminal_value": num(ws["B59"].value),
    "pv_terminal":    num(ws["B60"].value),
    "enterprise_value": num(ws["B65"].value),
    "pct_ev_terminal": num(ws["B67"].value),
    "cash":           num(ws["B72"].value),
    "total_debt_bridge": num(ws["B73"].value),
    "pref_bridge":    num(ws["B74"].value),
    "equity_value":   num(ws["B76"].value),
    "shares":         num(ws["B80"].value),
    "value_per_share": num(ws["B81"].value),
}

# ---------------------------------------------------------------- Sensitivity
ws = wb["Sensitivity Analysis"]
g_grid = [num(ws[f"{c}10"].value) for c in ["B", "C", "D", "E", "F"]]
sens = {}
for r in range(11, 16):
    wv = num(ws[f"A{r}"].value)
    sens[f"{wv:.4f}"] = {f"{gg:.4f}": num(ws[f"{c}{r}"].value)
                         for gg, c in zip(g_grid, ["B", "C", "D", "E", "F"])}
g["sensitivity"] = {"wacc_rows": [num(ws[f"A{r}"].value) for r in range(11, 16)],
                    "g_cols": g_grid, "grid": sens}

# ---------------------------------------------------------------- Debt schedule
ws = wb["Debt Schedule"]
g["debt"] = {
    "opening_tl":    row(ws, 6, ALL, YRS_SCHED),
    "tl_net_draw":   row(ws, 7, ALL, YRS_SCHED),
    "closing_tl":    row(ws, 8, ALL, YRS_SCHED),
    "tl_interest":   row(ws, 9, ALL, YRS_SCHED),
    "opening_bc":    row(ws, 12, ALL, YRS_SCHED),
    "bc_net_draw":   row(ws, 13, ALL, YRS_SCHED),
    "closing_bc":    row(ws, 14, ALL, YRS_SCHED),
    "bc_interest":   row(ws, 15, ALL, YRS_SCHED),
    "wc_outstanding": row(ws, 18, ALL, YRS_SCHED),
    "wc_interest":   row(ws, 19, ALL, YRS_SCHED),
    "bank_charges":  row(ws, 21, ALL, YRS_SCHED),
    "total_debt":    row(ws, 24, ALL, YRS_SCHED),
    "total_interest": row(ws, 25, ALL, YRS_SCHED),
}

# ---------------------------------------------------------------- Depreciation
ws = wb["Depreciation Schedule"]
g["depreciation"] = {
    "opening_gross_block": row(ws, 11, PROJ, YRS_SCHED),
    "additions_capex":     row(ws, 13, PROJ, YRS_SCHED),
    "book_depreciation":   row(ws, 21, PROJ, YRS_SCHED),
    "it_depreciation":     row(ws, 31, PROJ, YRS_SCHED),
}

# ---------------------------------------------------------------- DSCR / Covenant
# DSCR sheet layout: C=FY25A D=FY26E ... K=FY33E  (same as YRS_3S for proj years)
ws = wb["DSCR Covenant Tracker"]
g["covenant"] = {
    "pat":               row(ws, 7, PROJ, YRS_3S),
    "depreciation":      row(ws, 8, PROJ, YRS_3S),
    "interest_grossup":  row(ws, 9, PROJ, YRS_3S),
    "deferred_tax":      row(ws, 10, PROJ, YRS_3S),
    "cads":              row(ws, 11, PROJ, YRS_3S),
    "interest_paid":     row(ws, 14, PROJ, YRS_3S),
    "principal_repay":   row(ws, 15, PROJ, YRS_3S),
    "total_debt_service": row(ws, 16, PROJ, YRS_3S),
    "dscr":              row(ws, 19, PROJ, YRS_3S),
    "iscr":              row(ws, 20, PROJ, YRS_3S),
    "facr":              row(ws, 21, PROJ, YRS_3S),
    "debt_ebitda":       row(ws, 22, PROJ, YRS_3S),
    "net_debt_tnw":      row(ws, 23, PROJ, YRS_3S),
}

# ---------------------------------------------------------------- Revenue build
ws = wb["Assumptions & Drivers"]
# Section 6 columns: D=FY25A E=FY26E ... L=FY33E
YRS_AD = {"FY25A": "D", "FY26E": "E", "FY27E": "F", "FY28E": "G", "FY29E": "H",
          "FY30E": "I", "FY31E": "J", "FY32E": "K", "FY33E": "L"}
g["revenue"] = {
    "cob1_production": row(ws, 85, PROJ, YRS_AD),
    "cob2_production": row(ws, 88, PROJ, YRS_AD),
    "total_coke":      row(ws, 89, PROJ, YRS_AD),
    "coke_revenue":    row(ws, 91, PROJ, YRS_AD),
    "cog_revenue":     row(ws, 95, PROJ, YRS_AD),
    "tar_revenue":     row(ws, 99, PROJ, YRS_AD),
    "total_net_sales": row(ws, 100, PROJ, YRS_AD),
    "cdq_saving":      row(ws, 102, PROJ, YRS_AD),
}

# ---------------------------------------------------------------- Scenario outputs
ws = wb["Scenario Engine"]
g["scenario_engine"] = {
    "active": ws["B3"].value,
    "wacc": num(ws["B36"].value),
    "terminal_g": num(ws["B37"].value),
    "net_sales_fy25a": num(ws["B38"].value),
    "implied_ev": num(ws["D36"].value),
    "equity_value": num(ws["D37"].value),
    "vps": num(ws["D38"].value),
}

# ---------------------------------------------------------------- Ratio Analysis
ws = wb["Ratio Analysis"]
# Ratio Analysis projection columns: N=FY26E O=FY27E ... U=FY33E
YRS_RA = {"FY26E": "N", "FY27E": "O", "FY28E": "P", "FY29E": "Q", "FY30E": "R",
          "FY31E": "S", "FY32E": "T", "FY33E": "U"}
PROJ_E = list(YRS_RA.keys())
g["ratio_analysis"] = {
    "gross_margin":      row(ws, 12, PROJ_E, YRS_RA),
    "ebitda_margin":     row(ws, 13, PROJ_E, YRS_RA),
    "net_profit_margin": row(ws, 16, PROJ_E, YRS_RA),
    "current_ratio":     row(ws, 23, PROJ_E, YRS_RA),
    "debt_equity":       row(ws, 25, PROJ_E, YRS_RA),
    "ebit_int_cover":    row(ws, 26, PROJ_E, YRS_RA),
    "roce":              row(ws, 30, PROJ_E, YRS_RA),
    "roe":               row(ws, 32, PROJ_E, YRS_RA),
}

with open("golden.json", "w") as f:
    json.dump(g, f, indent=1)

print("golden.json written. Key Base outputs:")
print(f"  WACC = {g['dcf']['wacc']:.6f}")
print(f"  EV   = {g['dcf']['enterprise_value']:.4f}")
print(f"  VPS  = {g['dcf']['value_per_share']:.4f}")
print(f"  TV   = {g['dcf']['terminal_value']:.4f}")
print(f"  min DSCR = {min(v for v in g['covenant']['dscr'].values() if v):.4f}")
