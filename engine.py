"""
engine.py — Jindal Coke Ltd. Financial Engine (Excel-calibrated)
================================================================
Pure Python/Pandas; no Streamlit, no network. All currency in INR Crores.

This engine is a faithful digital twin of JCL_Financial_Model_EXP.xlsx. The
exogenous schedules the workbook computes on dedicated sheets (depreciation,
the term-loan / buyers-credit interest cascade, the debt & covenant schedules,
FY24A/FY25A actuals, balance-sheet roll-forwards, the dividend/buy-back
profile) are loaded verbatim from the bundled Base model
(`jcl_base_model.json`, see model_data.py). The genuinely driver-sensitive
lines (production & revenue, COGS%, EBITDA, WACC, FCFF, the DCF and the
covenant ratios) are computed by formula on top of those schedules.

Consequence: at the Base preset every figure reproduces the workbook to full
precision; moving any slider perturbs the model continuously from that exact
base. Verified cell-by-cell against the source workbook by verify.py.
"""

from __future__ import annotations

import math
from dataclasses import dataclass, field
from datetime import date as _date
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

from model_data import (
    ALL_YEARS, HISTORICAL_YEAR, PROJECTION_YEARS, VALUED_YEARS,
    base_model, scalar, series,
)

# =============================================================================
# CONSTANTS — sourced from JCL_Financial_Model_EXP.xlsx
# =============================================================================

COB1_CAPACITY_MT = 429_000
COB2_CAPACITY_MT = 350_000
COB1_UTIL_FLAT   = 0.911766

COG_YIELD_PER_MT_COKE = 222
COG_PRICE_INR_NM3     = 19.44
COAL_TAR_YIELD_PCT    = 0.0378
COAL_TAR_PRICE_INR_MT = 37_000

EMP_PCT_SALES = 0.0093
SGA_PCT_SALES = 0.0374

SHARES_OUTSTANDING_CR = 3.243
PREF_SHARE_CAPITAL    = 109.26
EFFECTIVE_TAX_RATE    = 0.2517
MAT_TAX_RATE          = 0.17472

# Capital-structure book values feeding the WACC weights (DCF sheet B13:B16)
WACC_EQUITY = 797.60
WACC_DEBT   = 545.82
WACC_PREF   = PREF_SHARE_CAPITAL

# EV→equity bridge (DCF sheet) — FY25A ending cash & total debt
BRIDGE_CASH = 210.18
BRIDGE_DEBT = 545.82

# Base anchors for sliders that perturb embedded schedules
BASE_COGS_ANCHOR = 0.82
BASE_INTEREST    = 0.09
BASE_CAPEX_INT   = 0.025
BASE_COB2_STEADY = 0.80

# Per-year Base COGS% profile; the cogs_pct slider shifts it by (slider − anchor)
COGS_PCT_BASE = {"FY26E": 0.79, "FY27E": 0.81, "FY28E": 0.808, "FY29E": 0.806,
                 "FY30E": 0.806, "FY31E": 0.806, "FY32E": 0.806, "FY33E": 0.806}

# Working-capital days — reproduce the workbook's WC levels exactly
DEBTOR_DAYS    = {"FY26E": 34, "FY27E": 33, "FY28E": 32, "FY29E": 31,
                  "FY30E": 31, "FY31E": 31, "FY32E": 31, "FY33E": 31}
INVENTORY_DAYS = {"FY26E": 43, "FY27E": 41, "FY28E": 40, "FY29E": 40,
                  "FY30E": 40, "FY31E": 40, "FY32E": 40, "FY33E": 40}
PAYABLE_DAYS   = {y: 40 for y in VALUED_YEARS}

# Sensitivity grid axes — match 'Sensitivity Analysis' sheet exactly
SENS_WACC_ROWS = [0.1151, 0.1201, 0.1251, 0.1301, 0.1351]
SENS_G_COLS    = [0.0, 0.01, 0.02, 0.03, 0.04]

# =============================================================================
# SCENARIO PRESETS  (Scenario Engine: Bull / Base / Bear columns)
# =============================================================================

SCENARIO_PRESETS: Dict[str, Dict] = {
    "Base": {
        "cob2_util_steady": 0.80, "coke_realization": 27_000, "cogs_pct": 0.82,
        "interest_rate": 0.09, "target_de": 0.82, "unlevered_beta": 0.85,
        "terminal_growth": 0.00, "capex_intensity": 0.025,
        "rf_rate": 0.07, "erp": 0.0725,
    },
    "Bull": {
        "cob2_util_steady": 0.85, "coke_realization": 30_000, "cogs_pct": 0.79,
        "interest_rate": 0.085, "target_de": 0.60, "unlevered_beta": 0.80,
        "terminal_growth": 0.02, "capex_intensity": 0.020,
        "rf_rate": 0.065, "erp": 0.07,
    },
    "Bear": {
        "cob2_util_steady": 0.75, "coke_realization": 24_000, "cogs_pct": 0.85,
        "interest_rate": 0.10, "target_de": 1.00, "unlevered_beta": 0.95,
        "terminal_growth": 0.00, "capex_intensity": 0.030,
        "rf_rate": 0.075, "erp": 0.075,
    },
}

INCOME_COLS = ["Net_Sales", "COGS", "Gross_Profit", "Employee", "SGA", "EBITDA",
               "Depreciation", "EBIT", "Other_Income", "Interest", "PBT", "Tax",
               "PAT", "Capex", "Opening_Gross_Block", "Closing_Term_Debt"]


def _is_base(a: Dict) -> bool:
    base = SCENARIO_PRESETS["Base"]
    return all(abs(float(a.get(k, base[k])) - base[k]) < 1e-9 for k in base)


# =============================================================================
# THE ENGINE
# =============================================================================

@dataclass
class JCLFinancialEngine:
    """3-statement + DCF + WACC engine for Jindal Coke Ltd."""
    assumptions: Dict = field(default_factory=lambda: dict(SCENARIO_PRESETS["Base"]))

    # ---- public API --------------------------------------------------------
    def build(self) -> Dict:
        revenue_df = self._build_revenue()
        income_df  = self._build_income_statement(revenue_df)
        bs_df, cfs_df = self._build_balance_sheet_and_cashflow(income_df)
        ratios_df  = self._build_ratios(income_df, bs_df, cfs_df)
        dcf        = self._build_dcf(income_df, cfs_df)
        sens       = self._build_sensitivity_grid(income_df, cfs_df, dcf)
        return {
            "revenue": revenue_df, "income": income_df, "balance": bs_df,
            "cashflow": cfs_df, "ratios": ratios_df, "dcf": dcf,
            "sensitivity": sens, "assumptions": self.assumptions.copy(),
        }

    # ---- COB-2 utilisation profile (driver: cob2_util_steady) --------------
    def _cob2_util(self) -> Dict[str, float]:
        s = self.assumptions["cob2_util_steady"]
        prof = {"FY25A": min(0.50, s * (0.35 / BASE_COB2_STEADY)),
                "FY26E": min(0.85, s * (0.70 / BASE_COB2_STEADY))}
        for y in VALUED_YEARS[1:]:  # FY27E onward
            prof[y] = s
        prof["FY26E"] = min(0.85, s * (0.70 / BASE_COB2_STEADY))
        prof["FY27E"] = s
        return prof

    # ---- revenue build (production × prices) -------------------------------
    def _build_revenue(self) -> pd.DataFrame:
        a = self.assumptions
        cob2_util = self._cob2_util()
        rows = []
        for y in PROJECTION_YEARS:
            cob1 = COB1_CAPACITY_MT * COB1_UTIL_FLAT
            cob2 = COB2_CAPACITY_MT * cob2_util[y]
            coke = cob1 + cob2
            coke_rev = coke * a["coke_realization"] / 1e7
            cog_rev = coke * COG_YIELD_PER_MT_COKE * COG_PRICE_INR_NM3 / 1e7
            tar_rev = coke * COAL_TAR_YIELD_PCT * COAL_TAR_PRICE_INR_MT / 1e7
            rows.append({
                "Year": y, "COB1_Production": cob1, "COB2_Production": cob2,
                "Total_Coke": coke, "Coke_Revenue": coke_rev,
                "COG_Revenue": cog_rev, "Tar_Revenue": tar_rev,
                "Total_Revenue": coke_rev + cog_rev + tar_rev,
                "Utilization": coke / (COB1_CAPACITY_MT + COB2_CAPACITY_MT),
            })
        return pd.DataFrame(rows).set_index("Year")

    # ---- income statement --------------------------------------------------
    def _build_income_statement(self, revenue_df: pd.DataFrame) -> pd.DataFrame:
        a = self.assumptions
        gi = base_model()["income"]
        glb = base_model()["balance"]
        dep_base   = series("depreciation", "book_depreciation")
        cdq        = series("revenue", "cdq_saving")
        int_rate_mult = a["interest_rate"] / BASE_INTEREST

        cogs_shift = a["cogs_pct"] - BASE_COGS_ANCHOR
        rows = []

        # FY24A & FY25A — actuals embedded verbatim
        for y in (HISTORICAL_YEAR, "FY25A"):
            rows.append({
                "Year": y,
                "Net_Sales": gi["Net_Sales"][y], "COGS": gi["COGS"][y],
                "Gross_Profit": gi["Gross_Profit"][y], "Employee": gi["Employee"][y],
                "SGA": gi["SGA"][y], "EBITDA": gi["EBITDA"][y],
                "Depreciation": gi["Depreciation"][y], "EBIT": gi["EBIT"][y],
                "Other_Income": gi["Other_Income"][y], "Interest": gi["Interest"][y],
                "PBT": gi["PBT"][y], "Tax": gi["Tax"][y], "PAT": gi["PAT"][y],
                "Capex": series("depreciation", "additions_capex").get(y, 0.0)
                         if y != HISTORICAL_YEAR else 0.0,
                "Opening_Gross_Block": series("depreciation", "opening_gross_block").get(y, 0.0)
                         if y != HISTORICAL_YEAR else 0.0,
                "Closing_Term_Debt": glb["LT_Borrowings"][y],
            })

        ogb = series("depreciation", "opening_gross_block")
        for y in VALUED_YEARS:
            sales = revenue_df.loc[y, "Total_Revenue"]
            cogs_pct = COGS_PCT_BASE[y] + cogs_shift
            cogs = sales * cogs_pct - cdq[y]
            gp = sales - cogs
            emp = sales * EMP_PCT_SALES
            sga = sales * SGA_PCT_SALES
            ebitda = sales - cogs - emp - sga
            dep = dep_base[y]
            ebit = ebitda - dep
            other_inc = gi["Other_Income"][y]
            interest = gi["Interest"][y] * int_rate_mult
            pbt = ebit + other_inc - interest
            tax = max(pbt, 0.0) * EFFECTIVE_TAX_RATE
            pat = pbt - tax
            capex = ogb[y] * a["capex_intensity"]
            rows.append({
                "Year": y, "Net_Sales": sales, "COGS": cogs, "Gross_Profit": gp,
                "Employee": emp, "SGA": sga, "EBITDA": ebitda, "Depreciation": dep,
                "EBIT": ebit, "Other_Income": other_inc, "Interest": interest,
                "PBT": pbt, "Tax": tax, "PAT": pat, "Capex": capex,
                "Opening_Gross_Block": ogb[y], "Closing_Term_Debt": glb["LT_Borrowings"][y],
            })

        df = pd.DataFrame(rows).set_index("Year")
        df["EBITDA_Margin"] = df["EBITDA"] / df["Net_Sales"]
        df["PAT_Margin"]    = df["PAT"] / df["Net_Sales"]
        df["Gross_Margin"]  = df["Gross_Profit"] / df["Net_Sales"]
        return df

    # ---- working-capital levels (days-based) -------------------------------
    def _wc_levels(self, income_df: pd.DataFrame) -> Dict[str, Dict[str, float]]:
        gb = base_model()["balance"]
        inv = {"FY25A": gb["Inventories"]["FY25A"]}
        ar  = {"FY25A": gb["Trade_Receivables"]["FY25A"]}
        ap  = {"FY25A": gb["Trade_Payables"]["FY25A"]}
        for y in VALUED_YEARS:
            sales = income_df.loc[y, "Net_Sales"]
            cogs = income_df.loc[y, "COGS"]
            inv[y] = cogs / 365 * INVENTORY_DAYS[y]
            ar[y] = sales / 365 * DEBTOR_DAYS[y]
            ap[y] = cogs / 365 * PAYABLE_DAYS[y]
        return {"inv": inv, "ar": ar, "ap": ap}

    # ---- balance sheet + cash flow -----------------------------------------
    def _build_balance_sheet_and_cashflow(
        self, income_df: pd.DataFrame
    ) -> Tuple[pd.DataFrame, pd.DataFrame]:
        gb = base_model()["balance"]
        gc = base_model()["cashflow"]
        wc = self._wc_levels(income_df)

        bs_rows, cfs_rows = [], []

        # FY24A & FY25A balance sheet — actuals
        for y in (HISTORICAL_YEAR, "FY25A"):
            bs_rows.append(self._bs_row_from_golden(gb, y))

        # FY25A cash flow — actual
        cfs_rows.append(self._cfs_row_from_golden(gc, "FY25A"))

        # roll-forward state
        reserves = gb["Reserves_Surplus"]["FY25A"]
        share_cap = gb["Share_Capital"]["FY25A"]
        cash = gb["Cash"]["FY25A"]
        base_pat = base_model()["income"]["PAT"]

        for i, y in enumerate(VALUED_YEARS):
            prev = VALUED_YEARS[i - 1] if i > 0 else "FY25A"
            r = income_df.loc[y]
            pat = r["PAT"]; dep = r["Depreciation"]; capex = r["Capex"]

            inv, ar, ap = wc["inv"][y], wc["ar"][y], wc["ap"][y]
            d_inv = wc["inv"][prev] - inv
            d_ar  = wc["ar"][prev] - ar
            d_ap  = ap - wc["ap"][prev]
            d_other_ncl = gc["Delta_Other_NCL"][y]
            cfo = pat + dep + d_inv + d_ar + d_ap + d_other_ncl

            # CFI: maintenance capex + embedded LT-loan/other realisation
            cfi_other = gc["CFI"][y] - gc["Capex"][y]   # non-capex part of CFI (Base)
            cfi = -capex + cfi_other

            # CFF: term-debt + ST-borrowing schedule embedded; dividend flexes on PAT
            div_base = gc["Dividend"][y]
            dividend = div_base - 0.30 * (pat - base_pat[y])   # variable payout flex
            cff = gc["Term_Debt_Repay"][y] + gc["Change_ST_Borrow"][y] + dividend

            revolver_draw = gc["Revolver_Draw"][y]
            net_change = cfo + cfi + cff
            opening_cash = cash
            cash = opening_cash + net_change + revolver_draw

            cfs_rows.append({
                "Year": y, "PAT": pat, "Depreciation": dep,
                "Delta_Inventory": d_inv, "Delta_Receivables": d_ar,
                "Delta_Payables": d_ap, "Delta_Other_NCL": d_other_ncl, "CFO": cfo,
                "Capex": -capex, "CFI": cfi,
                "Term_Debt_Repay": gc["Term_Debt_Repay"][y],
                "Change_ST_Borrow": gc["Change_ST_Borrow"][y],
                "Dividend": dividend, "CFF": cff,
                "Revolver_Draw": revolver_draw,
                "Net_Change_Cash": net_change, "Closing_Cash": cash,
            })

            # balance sheet — schedule items embedded; equity rolls on PAT
            reserves = reserves + pat + dividend  # dividend already negative
            total_equity = share_cap + reserves
            ltb = gb["LT_Borrowings"][y]; stb = gb["ST_Borrowings"][y]
            rev = gb["Revolver"][y]
            dtl = gb["Deferred_Tax_Liab"][y]; prov = gb["LT_Provisions"][y]
            other_cl = gb["Other_CL"][y]
            nfa = gb["Net_Fixed_Assets"][y]; cwip = gb["CWIP"][y]
            lt_inv = gb["LT_Investments"][y]; lt_loans = gb["LT_Loans_Adv"][y]
            other_lt = gb["Other_LT_Assets"][y]; other_ca = gb["Other_CA"][y]
            st_loans = gb["ST_Loans_Adv"][y]

            total_assets = (nfa + cwip + lt_inv + lt_loans + other_lt
                            + inv + ar + cash + other_ca + st_loans)
            total_le = (total_equity + ltb + dtl + prov + stb + rev + ap
                        + other_cl)
            bs_rows.append({
                "Year": y, "Share_Capital": share_cap,
                "Reserves_Surplus": reserves, "Total_Equity": total_equity,
                "LT_Borrowings": ltb, "Deferred_Tax_Liab": dtl,
                "LT_Provisions": prov, "ST_Borrowings": stb, "Revolver": rev,
                "Trade_Payables": ap, "Other_CL": other_cl,
                "Net_Fixed_Assets": nfa, "CWIP": cwip, "LT_Investments": lt_inv,
                "LT_Loans_Adv": lt_loans, "Other_LT_Assets": other_lt,
                "Inventories": inv, "Trade_Receivables": ar, "Cash": cash,
                "Other_CA": other_ca, "ST_Loans_Adv": st_loans,
                "Total_Assets": total_assets, "Total_Liab_Equity": total_le,
            })

        return (pd.DataFrame(bs_rows).set_index("Year"),
                pd.DataFrame(cfs_rows).set_index("Year"))

    @staticmethod
    def _bs_row_from_golden(gb: Dict, y: str) -> Dict:
        row = {"Year": y}
        for col in ["Share_Capital", "Reserves_Surplus", "Total_Equity",
                    "LT_Borrowings", "Deferred_Tax_Liab", "LT_Provisions",
                    "ST_Borrowings", "Revolver", "Trade_Payables", "Other_CL",
                    "Net_Fixed_Assets", "CWIP", "LT_Investments", "LT_Loans_Adv",
                    "Other_LT_Assets", "Inventories", "Trade_Receivables", "Cash",
                    "Other_CA", "ST_Loans_Adv", "Total_Assets", "Total_Liab_Equity"]:
            row[col] = gb[col][y]
        return row

    @staticmethod
    def _cfs_row_from_golden(gc: Dict, y: str) -> Dict:
        return {
            "Year": y, "PAT": gc["PAT"][y], "Depreciation": gc["Depreciation"][y],
            "Delta_Inventory": gc["Delta_Inventory"][y],
            "Delta_Receivables": gc["Delta_Receivables"][y],
            "Delta_Payables": gc["Delta_Payables"][y],
            "Delta_Other_NCL": gc["Delta_Other_NCL"][y], "CFO": gc["CFO"][y],
            "Capex": gc["Capex"][y], "CFI": gc["CFI"][y],
            "Term_Debt_Repay": gc["Term_Debt_Repay"][y],
            "Change_ST_Borrow": gc["Change_ST_Borrow"][y],
            "Dividend": gc["Dividend"][y], "CFF": gc["CFF"][y],
            "Revolver_Draw": gc["Revolver_Draw"][y],
            "Net_Change_Cash": gc["Net_Change_Cash"][y],
            "Closing_Cash": gc["Closing_Cash"][y],
        }

    # ---- covenant ratios (lender-style, matches DSCR Covenant Tracker) -----
    def _build_ratios(self, income_df, bs_df, cfs_df) -> pd.DataFrame:
        gov = base_model()["covenant"]
        rows = []
        for y in PROJECTION_YEARS:
            inc = income_df.loc[y]; bs = bs_df.loc[y]
            ebitda = inc["EBITDA"]; ebit = inc["EBIT"]; interest = inc["Interest"]
            pat = inc["PAT"]; dep = inc["Depreciation"]

            principal = gov["principal_repay"][y]
            deferred_tax = gov["deferred_tax"][y]
            cads = pat + dep + interest + deferred_tax
            debt_service = interest + principal
            dscr = cads / debt_service if debt_service else np.nan

            total_debt = bs["LT_Borrowings"] + bs["ST_Borrowings"] + bs["Revolver"]
            net_debt = total_debt - bs["Cash"]
            equity = bs["Total_Equity"]
            non_curr_liab = (bs["LT_Borrowings"] + bs["Deferred_Tax_Liab"]
                             + bs["LT_Provisions"])
            tca = (bs["Inventories"] + bs["Trade_Receivables"] + bs["Cash"]
                   + bs["Other_CA"] + bs["ST_Loans_Adv"])
            tcl = (bs["ST_Borrowings"] + bs["Revolver"] + bs["Trade_Payables"]
                   + bs["Other_CL"])

            rows.append({
                "Year": y,
                "EBITDA_Margin": inc["EBITDA_Margin"], "PAT_Margin": inc["PAT_Margin"],
                "ROE": pat / equity if equity else np.nan,
                "ROCE": ebit / (equity + non_curr_liab) if (equity + non_curr_liab) else np.nan,
                "Debt_Equity": total_debt / equity if equity else np.nan,
                "Net_Debt": net_debt,
                "Net_Debt_EBITDA": net_debt / ebitda if ebitda else np.nan,
                "DSCR": dscr,
                "Interest_Coverage": ebitda / interest if interest else np.nan,
                "Current_Ratio": tca / tcl if tcl else np.nan,
                "Total_Debt": total_debt,
                "Gross_Margin": inc["Gross_Margin"],
            })
        return pd.DataFrame(rows).set_index("Year")

    # ---- WACC --------------------------------------------------------------
    def _compute_wacc(self) -> Dict:
        a = self.assumptions
        de = a["target_de"]
        beta_l = a["unlevered_beta"] * (1 + (1 - EFFECTIVE_TAX_RATE) * de)
        ke = a["rf_rate"] + beta_l * a["erp"]
        kd_at = a["interest_rate"] * (1 - EFFECTIVE_TAX_RATE)
        kp = 0.09  # cost of preference capital (Scenario Engine, Base)
        # Weights from target D/E (WACC & DCF sheet B29:B31), debt/pref split fixed
        debt_frac = WACC_DEBT / (WACC_DEBT + WACC_PREF)
        pref_frac = WACC_PREF / (WACC_DEBT + WACC_PREF)
        we = 1.0 / (1.0 + de)
        wd = (de / (1.0 + de)) * debt_frac
        wp = (de / (1.0 + de)) * pref_frac
        wacc = we * ke + wd * kd_at + wp * kp
        return {"WACC": wacc, "Ke": ke, "Kd_AT": kd_at, "Kp": kp,
                "Beta_Levered": beta_l, "We": we, "Wd": wd, "Wp": wp}

    # ---- DCF ---------------------------------------------------------------
    def _build_dcf(self, income_df, cfs_df) -> Dict:
        wacc_pkg = self._compute_wacc()
        wacc = wacc_pkg["WACC"]
        g = self.assumptions["terminal_growth"]

        fcff_rows = []
        for i, y in enumerate(VALUED_YEARS, start=1):
            inc = income_df.loc[y]
            nopat = inc["EBIT"] * (1 - MAT_TAX_RATE)
            d_nwc = (cfs_df.loc[y, "Delta_Inventory"]
                     + cfs_df.loc[y, "Delta_Receivables"]
                     + cfs_df.loc[y, "Delta_Payables"])
            fcff = nopat + inc["Depreciation"] - inc["Capex"] + d_nwc
            disc = (1 + wacc) ** i
            fcff_rows.append({
                "Year": y, "T": i, "EBIT": inc["EBIT"], "NOPAT": nopat,
                "Depreciation": inc["Depreciation"], "Capex": inc["Capex"],
                "Delta_NWC": d_nwc, "FCFF": fcff,
                "Discount_Factor": disc, "PV_FCFF": fcff / disc,
            })
        fcff_df = pd.DataFrame(fcff_rows).set_index("Year")

        terminal_fcff_cit = (income_df.loc["FY33E", "EBIT"] * (1 - EFFECTIVE_TAX_RATE)
                             + income_df.loc["FY33E", "Depreciation"]
                             - income_df.loc["FY33E", "Capex"])
        terminal_value = terminal_fcff_cit * (1 + g) / (wacc - g) if wacc > g else np.nan
        pv_terminal = (terminal_value / ((1 + wacc) ** len(VALUED_YEARS))
                       if not np.isnan(terminal_value) else np.nan)
        sum_pv_fcff = fcff_df["PV_FCFF"].sum()
        ev = sum_pv_fcff + pv_terminal
        equity_value = ev + BRIDGE_CASH - BRIDGE_DEBT - PREF_SHARE_CAPITAL
        vps = equity_value / SHARES_OUTSTANDING_CR
        pct_ev_terminal = pv_terminal / ev if (ev and not np.isnan(ev)) else np.nan
        return {
            "wacc_components": wacc_pkg, "fcff": fcff_df,
            "terminal_fcff_cit": terminal_fcff_cit,
            "terminal_value": terminal_value, "pv_terminal": pv_terminal,
            "sum_pv_fcff": sum_pv_fcff, "enterprise_value": ev,
            "cash": BRIDGE_CASH, "debt": BRIDGE_DEBT,
            "preference": PREF_SHARE_CAPITAL, "equity_value": equity_value,
            "value_per_share": vps, "pct_ev_terminal": pct_ev_terminal,
            "wacc": wacc, "terminal_growth": g,
        }

    # ---- sensitivity grid (matches 'Sensitivity Analysis' sheet) -----------
    def _build_sensitivity_grid(self, income_df, cfs_df, dcf) -> pd.DataFrame:
        fcff = [dcf["fcff"].loc[y, "FCFF"] for y in VALUED_YEARS]
        terminal_fcff_cit = dcf["terminal_fcff_cit"]
        base_wacc = dcf["wacc"]
        n = len(fcff)
        # Excel rows = base WACC + {-0.01, -0.005, 0, +0.005, +0.01}; cols = g grid
        wacc_rows = [base_wacc + d for d in (-0.01, -0.005, 0.0, 0.005, 0.01)]
        grid = np.full((len(wacc_rows), len(SENS_G_COLS)), np.nan)
        for iw, w in enumerate(wacc_rows):
            for jg, gg in enumerate(SENS_G_COLS):
                if w <= gg:
                    continue
                pv_fcff = sum(f / ((1 + w) ** (k + 1)) for k, f in enumerate(fcff))
                tv = terminal_fcff_cit * (1 + gg) / (w - gg)
                pv_tv = tv / ((1 + w) ** n)
                ev = pv_fcff + pv_tv
                eqv = ev + BRIDGE_CASH - BRIDGE_DEBT - PREF_SHARE_CAPITAL
                grid[iw, jg] = eqv / SHARES_OUTSTANDING_CR
        return pd.DataFrame(grid, index=[f"{w:.2%}" for w in wacc_rows],
                            columns=[f"{g:.2%}" for g in SENS_G_COLS])

    # ---- tornado -----------------------------------------------------------
    def tornado_analysis(self) -> pd.DataFrame:
        baseline = self.build()
        base_vps = baseline["dcf"]["value_per_share"]
        drivers = {
            "Coke Realization": ("coke_realization", 0.10),
            "COGS %": ("cogs_pct", 0.05),
            "Capex Intensity": ("capex_intensity", 0.20),
            "Interest Rate": ("interest_rate", 0.10),
            "Unlevered Beta": ("unlevered_beta", 0.10),
            "Risk-Free Rate": ("rf_rate", 0.10),
            "COB-2 Utilization": ("cob2_util_steady", 0.10),
            "Target D/E": ("target_de", 0.20),
        }
        rows = []
        original = self.assumptions.copy()
        for label, (key, pct) in drivers.items():
            base_val = original[key]
            self.assumptions = original.copy(); self.assumptions[key] = base_val * (1 + pct)
            up_vps = self.build()["dcf"]["value_per_share"]
            self.assumptions = original.copy(); self.assumptions[key] = base_val * (1 - pct)
            dn_vps = self.build()["dcf"]["value_per_share"]
            rows.append({"Driver": label, "Down": dn_vps - base_vps,
                         "Up": up_vps - base_vps, "Range": abs(up_vps - dn_vps)})
        self.assumptions = original
        return pd.DataFrame(rows).sort_values("Range", ascending=True)

    # ---- monte carlo -------------------------------------------------------
    def monte_carlo(self, n: int = 1_000, seed: int = 42) -> np.ndarray:
        rng = np.random.default_rng(seed)
        original = self.assumptions.copy()
        evs = np.zeros(n)
        keys_vol = {"coke_realization": 0.10, "cogs_pct": 0.05,
                    "cob2_util_steady": 0.07, "interest_rate": 0.10,
                    "capex_intensity": 0.15}
        for i in range(n):
            self.assumptions = original.copy()
            for k, v in keys_vol.items():
                shock = rng.normal(0, v / 1.96)
                self.assumptions[k] = original[k] * (1 + shock)
            try:
                evs[i] = self.build()["dcf"]["enterprise_value"]
            except Exception:
                evs[i] = np.nan
        self.assumptions = original
        return evs[~np.isnan(evs)]


# =============================================================================
# INSIGHTS ENGINE
# =============================================================================

def generate_insights(results: Dict) -> List[Dict]:
    out: List[Dict] = []
    dcf, ratios, income = results["dcf"], results["ratios"], results["income"]

    pct_tv = dcf["pct_ev_terminal"]
    if pct_tv > 0.80:
        out.append({"level": "warning", "icon": "WARN",
                    "text": f"{pct_tv:.0%} of EV from Terminal Value - perpetuity assumptions dominate. Stress-test g and WACC."})
    elif pct_tv > 0.60:
        out.append({"level": "caution", "icon": "INFO",
                    "text": f"{pct_tv:.0%} of EV from Terminal Value - moderately dependent on long-run assumptions."})
    else:
        out.append({"level": "good", "icon": "OK",
                    "text": f"Healthy: only {pct_tv:.0%} of EV from Terminal Value - explicit-period cash flows do most of the lifting."})

    min_dscr = ratios["DSCR"].min(); min_dscr_year = ratios["DSCR"].idxmin()
    if min_dscr < 1.20:
        out.append({"level": "alert", "icon": "ALERT",
                    "text": f"DSCR drops to {min_dscr:.2f}x in {min_dscr_year} - covenant breach. Review debt-service capacity."})
    elif min_dscr < 1.50:
        out.append({"level": "caution", "icon": "WATCH",
                    "text": f"Min DSCR of {min_dscr:.2f}x in {min_dscr_year} - acceptable but tight."})
    else:
        out.append({"level": "good", "icon": "OK",
                    "text": f"Strong: min DSCR of {min_dscr:.2f}x in {min_dscr_year} - comfortable debt-service coverage."})

    max_nd = ratios["Net_Debt_EBITDA"].max()
    if max_nd > 3.0:
        out.append({"level": "warning", "icon": "WARN",
                    "text": f"Leverage warning: peak Net Debt/EBITDA of {max_nd:.2f}x - above prudent threshold."})
    elif max_nd > 2.0:
        out.append({"level": "caution", "icon": "INFO",
                    "text": f"Moderate leverage: peak Net Debt/EBITDA of {max_nd:.2f}x - manageable."})
    else:
        out.append({"level": "good", "icon": "OK",
                    "text": f"Deleveraged: peak Net Debt/EBITDA of {max_nd:.2f}x - strong balance sheet."})

    wacc = dcf["wacc"]
    if wacc < 0.10:
        out.append({"level": "warning", "icon": "WARN",
                    "text": f"Low WACC: {wacc:.2%} - verify capital-structure assumptions; appears aggressive."})
    elif wacc > 0.16:
        out.append({"level": "caution", "icon": "INFO",
                    "text": f"High WACC: {wacc:.2%} - conservative discount rate compresses valuation."})

    margin_25 = income.loc["FY25A", "EBITDA_Margin"]
    margin_29 = income.loc["FY29E", "EBITDA_Margin"]
    delta = margin_29 - margin_25
    if delta > 0.05:
        out.append({"level": "good", "icon": "OK",
                    "text": f"Margin expansion: EBITDA grows {delta * 100:.1f} pp from FY25A to FY29E - operating leverage from COB-2 ramp + CDQ savings."})
    elif delta < -0.02:
        out.append({"level": "warning", "icon": "WARN",
                    "text": f"Margin compression: EBITDA contracts {abs(delta) * 100:.1f} pp from FY25A to FY29E - cost pressure not offset by realization."})
    return out


# =============================================================================
# EXCEL SYNC — assumptions parser (kept for the sidebar 'pull assumptions')
# =============================================================================

def parse_excel_assumptions(file_buffer) -> dict:
    """Pull Base-column drivers from the Scenario Engine sheet. Returns {} on failure."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_buffer, data_only=True, read_only=True)
    except Exception:
        return {}
    candidate_sheets = ["Scenario Engine", "Scenario_Engine", "Scenarios",
                        "Drivers", "Assumptions"]
    sheet_name = next((s for s in candidate_sheets if s in wb.sheetnames), None)
    if sheet_name is None:
        try: wb.close()
        except Exception: pass
        return {}
    ws = wb[sheet_name]

    base_col = None
    for r in range(1, 30):
        for c in range(1, 8):
            try:
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and v.strip().lower() == "base":
                    neighbors = []
                    for dc in (-1, 1):
                        try:
                            nv = ws.cell(row=r, column=c + dc).value
                            if isinstance(nv, str):
                                neighbors.append(nv.strip().lower())
                        except Exception:
                            pass
                    if any(n in ("bull", "bear") for n in neighbors):
                        base_col = c; break
            except Exception:
                continue
        if base_col is not None:
            break
    if base_col is None:
        base_col = 3

    LABEL_MAP = [
        (["cogs (% of net sales)", "cogs %", "cogs % of net sales"], "cogs_pct", 0.50, 0.98),
        (["interest rate on debt", "interest rate (%)", "cost of debt"], "interest_rate", 0.04, 0.20),
        (["terminal growth rate"], "terminal_growth", -0.01, 0.08),
        (["unlevered beta"], "unlevered_beta", 0.40, 2.00),
        (["coke realization", "coke realisation"], "coke_realization", 10_000, 60_000),
        (["cob-2 util fy28", "cob2 util fy28", "steady-state util", "cob-2 utilization"], "cob2_util_steady", 0.20, 1.00),
        (["risk-free rate", "rf rate", "rf,", "risk free"], "rf_rate", 0.02, 0.15),
        (["equity risk premium", "erp"], "erp", 0.03, 0.15),
        (["target debt-to-equity", "target d/e"], "target_de", 0.10, 3.00),
    ]

    def _label(r):
        try:
            a = ws.cell(row=r, column=1).value; b = ws.cell(row=r, column=2).value
        except Exception:
            return ""
        return " | ".join(str(p).strip().lower() for p in (a, b) if p is not None)

    def _find(subs):
        for r in range(1, 201):
            lbl = _label(r)
            if not lbl:
                continue
            for needle in subs:
                if all(tok in lbl for tok in needle.split()):
                    return r
        return None

    def _read(r):
        for col in [base_col] + [c for c in (3, 4, 5) if c != base_col]:
            try:
                v = ws.cell(row=r, column=col).value
                if v is None:
                    continue
                return float(v)
            except (TypeError, ValueError):
                continue
        return None

    result: Dict[str, float] = {}
    for subs, key, lo, hi in LABEL_MAP:
        r = _find(subs)
        if r is None:
            continue
        v = _read(r)
        if v is None or not (lo <= v <= hi):
            continue
        result[key] = v
    try: wb.close()
    except Exception: pass
    return result


# =============================================================================
# REVERSE DCF SOLVERS
# =============================================================================

def solve_implied_beta(assumptions: dict, target_vps: float, lo: float = 0.30,
                       hi: float = 2.50, tol: float = 1e-4, max_iter: int = 60) -> Optional[Dict]:
    def _vps(beta):
        tmp = assumptions.copy(); tmp["unlevered_beta"] = beta
        try:
            return JCLFinancialEngine(assumptions=tmp).build()["dcf"]["value_per_share"]
        except Exception:
            return float("nan")
    lo_vps, hi_vps = _vps(lo), _vps(hi)
    if math.isnan(lo_vps) or math.isnan(hi_vps):
        return None
    if target_vps > lo_vps or target_vps < hi_vps:
        return None
    for _ in range(max_iter):
        mid = (lo + hi) / 2; mid_vps = _vps(mid)
        if math.isnan(mid_vps):
            return None
        if abs(mid_vps - target_vps) < tol * max(target_vps, 1.0):
            break
        if mid_vps > target_vps:
            lo = mid
        else:
            hi = mid
    final_beta = (lo + hi) / 2
    tmp = assumptions.copy(); tmp["unlevered_beta"] = final_beta
    try:
        res = JCLFinancialEngine(assumptions=tmp).build()
        return {"beta": final_beta, "implied_wacc": res["dcf"]["wacc"],
                "achieved_vps": res["dcf"]["value_per_share"]}
    except Exception:
        return None


def solve_implied_terminal_growth(assumptions: dict, target_vps: float, lo: float = -0.01,
                                  hi: float = 0.07, tol: float = 1e-4, max_iter: int = 60) -> Optional[Dict]:
    def _vps(g):
        tmp = assumptions.copy(); tmp["terminal_growth"] = g
        try:
            return JCLFinancialEngine(assumptions=tmp).build()["dcf"]["value_per_share"]
        except Exception:
            return float("nan")
    lo_vps, hi_vps = _vps(lo), _vps(hi)
    if math.isnan(lo_vps) or math.isnan(hi_vps):
        return None
    if target_vps < lo_vps or target_vps > hi_vps:
        return None
    for _ in range(max_iter):
        mid = (lo + hi) / 2; mid_vps = _vps(mid)
        if math.isnan(mid_vps):
            return None
        if abs(mid_vps - target_vps) < tol * max(target_vps, 1.0):
            break
        if mid_vps < target_vps:
            lo = mid
        else:
            hi = mid
    final_g = (lo + hi) / 2
    return {"terminal_growth": final_g, "achieved_vps": _vps(final_g)}


# =============================================================================
# COVENANT STRESS TESTER
# =============================================================================

def covenant_stress_sweep(assumptions: dict, driver_key: str = "interest_rate",
                          lo_override: Optional[float] = None, hi_override: Optional[float] = None,
                          steps: int = 40, covenant_floor: float = 1.20) -> dict:
    DEFAULT_RANGES = {
        "interest_rate": (0.04, 0.25),
        "cogs_pct": (0.70, 0.98),
        "cob2_util_steady": (0.20, max(0.20, assumptions.get("cob2_util_steady", 0.80))),
    }
    lo = lo_override if lo_override is not None else DEFAULT_RANGES.get(driver_key, (0.0, 1.0))[0]
    hi = hi_override if hi_override is not None else DEFAULT_RANGES.get(driver_key, (0.0, 1.0))[1]
    if hi <= lo:
        hi = lo + 0.05
    driver_vals = np.linspace(lo, hi, steps)
    min_dscr_series: List[float] = []
    breach_value = breach_year = None
    for v in driver_vals:
        tmp = assumptions.copy(); tmp[driver_key] = float(v)
        try:
            rat = JCLFinancialEngine(assumptions=tmp).build()["ratios"]
            min_d = rat["DSCR"].min(); min_y = rat["DSCR"].idxmin()
            min_dscr_series.append(float(min_d))
            if breach_value is None and min_d < covenant_floor:
                breach_value = float(v); breach_year = str(min_y)
        except Exception:
            min_dscr_series.append(float("nan"))
    return {"driver_values": [float(v) for v in driver_vals],
            "min_dscr_series": min_dscr_series, "breach_value": breach_value,
            "breach_year": breach_year, "driver_key": driver_key,
            "covenant_floor": covenant_floor}


# =============================================================================
# TEXT REPORT GENERATOR
# =============================================================================

def generate_text_report(results: dict, assumptions: dict, scenario: str) -> str:
    try:
        inc = results["income"]; dcf = results["dcf"]; rat = results["ratios"]
        wcc = dcf["wacc_components"]
    except Exception:
        return "# Report generation failed - results dict missing required keys.\n"

    def cr(v): return f"INR {v:,.1f} Cr"
    def pc(v): return f"{v * 100:.1f}%"

    sep = "-" * 60
    lines = [
        "# JINDAL COKE LTD - ANALYST REPORT",
        f"Scenario: {scenario.upper()}  |  Generated: {_date.today().isoformat()}",
        "Model: JCL EXP v2.0  |  Currency: INR Crores", "", sep,
        "## 1. VALUATION SUMMARY", sep,
        f"Enterprise Value         : {cr(dcf['enterprise_value'])}",
        f"(+) FY25A Cash           : {cr(dcf['cash'])}",
        f"(-) Total Debt           : {cr(dcf['debt'])}",
        f"(-) Preference Capital   : {cr(dcf['preference'])}",
        f"Equity Value             : {cr(dcf['equity_value'])}",
        "Shares Outstanding       : 3.243 Crore",
        f"Intrinsic Value/Share    : INR {dcf['value_per_share']:,.2f}", "",
        f"WACC                     : {pc(dcf['wacc'])}",
        f"Terminal Growth (g)      : {pc(dcf['terminal_growth'])}",
        f"% EV from Terminal Value : {dcf['pct_ev_terminal'] * 100:.1f}%", "",
        sep, "## 2. WACC BUILD", sep,
        f"Risk-free Rate (Rf)      : {pc(assumptions['rf_rate'])}",
        f"Equity Risk Premium      : {pc(assumptions['erp'])}",
        f"Unlevered Beta (bu)      : {assumptions['unlevered_beta']:.3f}",
        f"Relevered Beta (bL)      : {wcc['Beta_Levered']:.3f}",
        f"Cost of Equity (Ke)      : {pc(wcc['Ke'])}",
        f"After-tax Cost of Debt   : {pc(wcc['Kd_AT'])}",
        f"Cost of Preference (Kp)  : {pc(wcc['Kp'])}",
        f"Target D/E               : {assumptions['target_de']:.2f}x", "",
        sep, "## 3. INCOME STATEMENT (INR Cr)", sep,
        f"{'Year':<10}{'Revenue':>10}{'EBITDA':>10}{'EBITDA%':>9}{'PAT':>10}{'PAT%':>8}",
        "-" * 57,
    ]
    for y in inc.index:
        lines.append(
            f"{y:<10}{inc.loc[y, 'Net_Sales']:>10,.0f}{inc.loc[y, 'EBITDA']:>10,.0f}"
            f"{inc.loc[y, 'EBITDA_Margin'] * 100:>8.1f}%{inc.loc[y, 'PAT']:>10,.0f}"
            f"{inc.loc[y, 'PAT_Margin'] * 100:>7.1f}%")
    lines += ["", sep, "## 4. COVENANT METRICS", sep,
              f"{'Year':<10}{'DSCR':>8}{'ND/EBITDA':>12}{'Status':>12}", "-" * 42]
    for y in rat.index:
        d = rat.loc[y, "DSCR"]; nd = rat.loc[y, "Net_Debt_EBITDA"]
        flag = "PASS" if d >= 1.20 else "BREACH"
        lines.append(f"{y:<10}{d:>7.2f}x{nd:>11.2f}x{flag:>12}")
    lines += ["", sep, "## 5. KEY ASSUMPTIONS", sep,
              f"Coke Realization         : INR {assumptions['coke_realization']:,.0f}/MT",
              f"COB-2 Steady-state Util  : {pc(assumptions['cob2_util_steady'])}",
              f"COGS % (FY25A anchor)    : {pc(assumptions['cogs_pct'])}",
              f"Interest Rate on Debt    : {pc(assumptions['interest_rate'])}",
              f"Capex Intensity          : {pc(assumptions['capex_intensity'])}", "",
              sep, "## 6. RISK FLAGS", sep]
    try:
        for ins in generate_insights(results):
            lines.append(f"[{ins['icon']:5}] {ins['text']}")
    except Exception:
        lines.append("(Insights unavailable.)")
    lines += ["", sep,
              "DISCLAIMER: Illustrative model for institutional valuation training.",
              "Independent verification required for any investment decision.", sep]
    return "\n".join(lines)
